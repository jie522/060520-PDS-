#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
設備主檔匯入／索引重建工具

把 config.EQUIPMENT_CODING_XLSX 的「PDM新格式編碼」分頁匯入 equipment.db，
並掃描網路資料夾上的設備照片與技術資料，依 PDM 編碼前綴自動歸位。

設計重點（詳見 docs/equipment-master.md）：
  * equipment.db 是**正本**，Excel 只是匯入來源。匯入採合併（merge）而非整表重建：
    - source='excel' 的設備才會被 Excel 覆蓋
    - source='manual'（日後在系統內新增的）永遠不會被覆蓋或刪除
    - Excel 裡消失的設備不會被刪掉，只在報表裡提示
  * PDM 編碼（X##-###）是設備的主鍵，也是群組/類型/屬性/流水號的唯一真實來源；
    編碼解析不出來的（如 B-101）原樣匯入並標記 needs_fix，且不讓它污染編碼字典
  * 照片/技術資料表是從檔案系統推導出來的，每次掃描重建，但保留人工認領的對應關係

用法：
  python build_equipment_index.py              # 匯入 Excel + 掃描照片/技術資料
  python build_equipment_index.py --no-scan    # 只匯入 Excel，不掃網路資料夾（快）
  python build_equipment_index.py --deploy     # 完成後同步 equipment.db 至 dist_embed
"""

import os
import re
import sys
import shutil
import sqlite3
import datetime
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

_APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(_APP_DIR, 'equipment.db')

# PDM 編碼格式：群組(1碼) + 機台類型(2碼) + '-' + 加工屬性(1碼) + 流水號(2碼)
CODE_RE = re.compile(r'^([A-Z])(\d{2})-(\d)(\d{2})$')
# 資料夾名稱開頭帶編碼者可自動歸位，例如「A05-302同清」「B01-105(圓刀倒角)」
FOLDER_CODE_RE = re.compile(r'^([A-Z]\d{2}-\d{3})')
# 群組標題列，例如「 (A)  CNC車床/銑床類」
GROUP_TITLE_RE = re.compile(r'^\(([A-Z])\)\s*(.*)$')

PHOTO_EXT = {'.jpg', '.jpeg', '.png', '.bmp', '.gif', '.webp'}
SKIP_FILES = {'thumbs.db', 'desktop.ini'}

# Excel 欄位位置（0-based），對應「PDM新格式編碼」分頁的表頭
COL_GROUP, COL_TYPE, COL_TYPE_NAME = 0, 1, 2
COL_ATTR, COL_ATTR_NAME, COL_SEQ = 3, 4, 5
COL_OLD, COL_REMARK, COL_VENDOR, COL_CODE = 6, 7, 8, 9
COL_BUY, COL_NOTE, COL_MOVE, COL_LOCATION = 10, 11, 12, 13

SCHEMA = """
CREATE TABLE IF NOT EXISTS eq_group (
    code TEXT PRIMARY KEY, name TEXT, sort INTEGER
);
CREATE TABLE IF NOT EXISTS eq_type (
    group_code TEXT, code TEXT, name TEXT, sort INTEGER,
    PRIMARY KEY (group_code, code)
);
CREATE TABLE IF NOT EXISTS eq_attr (
    group_code TEXT, type_code TEXT, code TEXT, name TEXT,
    PRIMARY KEY (group_code, type_code, code)
);
CREATE TABLE IF NOT EXISTS equipment (
    code        TEXT PRIMARY KEY,
    group_code  TEXT, type_code TEXT, attr_code TEXT, seq TEXT,
    old_code    TEXT, vendor TEXT, model TEXT, buy_date TEXT,
    remark      TEXT, note TEXT,
    -- Excel 原始的類型/屬性名稱。編碼解析不出來的設備（needs_fix）不能信編碼字典
    -- （它的類型/屬性碼是硬湊的，查字典會查到別台設備的名稱），要顯示這裡的原文。
    type_name_raw TEXT, attr_name_raw TEXT,
    location    TEXT,
    status      TEXT DEFAULT '在用',
    -- source 與 origin 是兩件事，不可合併：
    --   source = 目前這筆能不能被 Excel 覆蓋（在系統內改過就變 manual）
    --   origin = 這筆資料當初打哪來的，寫入後永不變更，用來擋「Excel 來源的設備被硬刪」
    -- 曾經只用 source 一個欄位判斷，結果改過一次的 Excel 設備就變成可硬刪（2026-07 實測踩到）
    source      TEXT DEFAULT 'excel',
    origin      TEXT DEFAULT 'excel',
    needs_fix   INTEGER DEFAULT 0,
    fix_reason  TEXT,
    created_at  TEXT DEFAULT (datetime('now','localtime')),
    updated_at  TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS eq_spec (
    code TEXT, spec_name TEXT, spec_value TEXT, sort INTEGER
);
CREATE TABLE IF NOT EXISTS eq_photo (
    relpath TEXT PRIMARY KEY,             -- 相對 EQUIPMENT_PHOTO_ROOT 的路徑
    code    TEXT,                         -- NULL = 尚未歸位（資料夾名稱看不出編碼）
    folder  TEXT, filename TEXT,
    size    INTEGER, mtime TEXT,
    is_cover INTEGER DEFAULT 0,
    claimed INTEGER DEFAULT 0,            -- 1 = 人工認領的對應，重新掃描時保留
    sort    INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS eq_tech_file (
    relpath TEXT PRIMARY KEY,             -- 相對 EQUIPMENT_TECH_ROOT 的路徑
    code    TEXT, folder TEXT, filename TEXT,
    ext     TEXT, size INTEGER, mtime TEXT
);
CREATE TABLE IF NOT EXISTS eq_history (
    code TEXT, date TEXT, action TEXT, detail TEXT, user TEXT
);
CREATE INDEX IF NOT EXISTS idx_eq_group  ON equipment(group_code);
CREATE INDEX IF NOT EXISTS idx_eq_loc    ON equipment(location);
CREATE INDEX IF NOT EXISTS idx_eq_photo  ON eq_photo(code);
CREATE INDEX IF NOT EXISTS idx_eq_tech   ON eq_tech_file(code);
CREATE INDEX IF NOT EXISTS idx_eq_spec   ON eq_spec(code);
CREATE INDEX IF NOT EXISTS idx_eq_hist   ON eq_history(code);
"""


def _s(v):
    """儲存格轉字串：datetime 取日期，其餘去頭尾空白"""
    if v is None:
        return ''
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def parse_specs(remark):
    """把備註欄的規格文字拆成鍵值。

    來源長這樣（項目間可能是換行或緊接的「N.」）：
        1.床台:800mm*530mm
        2.刀庫30把
        3.X/Y/Z:36/36/30
        4.可中心出水
    拆法：先切項目 → 去掉開頭編號 → 有冒號就以冒號分名稱/值；
    沒冒號但「文字+數字」就從第一個數字切開（床台1150mm → 床台 / 1150mm）；
    純文字敘述（可中心出水）歸到「特性」。
    """
    if not remark:
        return []
    text = remark.replace('\r', '\n')
    parts = []
    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue
        # 同一行內連寫多個「N.」項目時再切一次
        parts.extend(p.strip() for p in re.split(r'(?=\d+\.)', line) if p.strip())

    specs = []
    for i, item in enumerate(parts):
        item = re.sub(r'^\d+\s*[.、]\s*', '', item).strip()
        if not item:
            continue
        m = re.match(r'^([^:：]+)[:：]\s*(.+)$', item)
        if m:
            name, value = m.group(1).strip(), m.group(2).strip()
        else:
            m2 = re.match(r'^([^\d]+?)\s*(\d.*)$', item)
            if m2:
                name, value = m2.group(1).strip(), m2.group(2).strip()
            else:
                name, value = '特性', item
        specs.append((name, value, i))
    return specs


def guess_status(note, remark):
    """從「說明」欄推斷設備狀態（賣掉/報廢/閒置），判斷不出來就當在用"""
    text = f'{note} {remark}'
    if '報廢' in text:
        return '報廢'
    if '賣' in text or '售' in text:
        return '已售出'
    if '閒置' in text:
        return '閒置'
    return '在用'


def parse_move_history(move_text):
    """把「設備移轉記錄」欄拆成歷程：2025/1/18移至詠設 → (2025-01-18, 移轉, 移至詠設)"""
    if not move_text:
        return []
    out = []
    for line in re.split(r'[\n;；]', move_text):
        line = line.strip()
        if not line:
            continue
        m = re.match(r'^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})\s*(.*)$', line)
        if m:
            date = f'{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
            detail = m.group(4).strip() or line
        else:
            date, detail = '', line
        out.append((date, '移轉', detail))
    return out


def _open_workbook(path):
    """讀取 Excel。網路檔案被鎖住時退而複製到暫存目錄再讀。"""
    import openpyxl
    try:
        return openpyxl.load_workbook(path, data_only=True, read_only=True)
    except (PermissionError, OSError) as e:
        print(f'  [WARN] 直接讀取失敗（{e}），改複製到暫存目錄再讀')
        tmp = os.path.join(tempfile.gettempdir(), '_eq_coding_tmp.xlsx')
        shutil.copy2(path, tmp)
        return openpyxl.load_workbook(tmp, data_only=True, read_only=True)


def import_excel(conn):
    """匯入 Excel 設備主檔，回傳 (新增, 更新, 略過manual, 待修正, 警告清單)"""
    path = config.EQUIPMENT_CODING_XLSX
    sheet = config.EQUIPMENT_CODING_SHEET
    print(f'讀取 Excel：{path}  [{sheet}]')
    if not os.path.exists(path):
        print(f'  [ERROR] 檔案不存在或無法存取：{path}')
        return None

    wb = _open_workbook(path)
    if sheet not in wb.sheetnames:
        print(f'  [ERROR] 找不到分頁「{sheet}」，現有分頁：{wb.sheetnames}')
        return None
    ws = wb[sheet]

    cur = conn.cursor()
    existing = {r[0]: r[1] for r in cur.execute('SELECT code, source FROM equipment')}

    # 合併儲存格會讓群組/類型/屬性只出現在第一列，往下要沿用上一個非空值（forward fill）；
    # 但層級變動時要清掉下層，避免上一組的屬性被錯誤帶到新的機台類型。
    ff_group = ff_type = ff_type_name = ff_attr = ff_attr_name = ''
    group_names = {}
    attr_owner = {}     # (group,type,attr) -> 第一個登記的名稱，後來的不同名稱視為撞號
    warnings = []
    added = updated = skipped = fixcnt = 0
    seen_codes = []

    for raw in ws.iter_rows(values_only=True):
        c = [_s(x) for x in raw] + [''] * 14
        first = c[COL_GROUP]

        # 群組標題列：「 (A)  CNC車床/銑床類」
        m = GROUP_TITLE_RE.match(first)
        if m:
            group_names[m.group(1)] = m.group(2).strip()
            ff_group = m.group(1)
            ff_type = ff_type_name = ff_attr = ff_attr_name = ''
            continue
        if first == '群組分類' or c[COL_CODE] == 'PDM編碼':
            continue                                    # 表頭列

        if first:
            if first != ff_group:
                ff_type = ff_type_name = ff_attr = ff_attr_name = ''
            ff_group = first
        if c[COL_TYPE]:
            if c[COL_TYPE] != ff_type:
                ff_attr = ff_attr_name = ''             # 換機台類型 → 屬性重新起算
            ff_type = c[COL_TYPE]
        if c[COL_TYPE_NAME]:
            ff_type_name = c[COL_TYPE_NAME]
        if c[COL_ATTR]:
            ff_attr = c[COL_ATTR]
        if c[COL_ATTR_NAME]:
            ff_attr_name = c[COL_ATTR_NAME]

        code = c[COL_CODE]
        if not code:
            continue                                    # 沒有 PDM 編碼的列（含尾端空列）不是設備

        # 編碼本身是群組/類型/屬性/流水號的真實來源，解析得出就以它為準
        fix_reasons = []
        mc = CODE_RE.match(code)
        if mc:
            group_code, type_code, attr_code, seq = mc.groups()
            if (group_code, type_code, attr_code, seq) != (ff_group, ff_type, ff_attr, c[COL_SEQ] or seq):
                pass    # 欄位與編碼不一致時以編碼為準，不另外報錯（欄位常有合併儲存格殘留）
        else:
            group_code, type_code = ff_group, ff_type
            attr_code, seq = ff_attr, c[COL_SEQ]
            fix_reasons.append(f'PDM編碼「{code}」不符 X##-### 格式')

        # 編碼字典：只採信格式正確的設備，避免 B-101 之類的錯誤編碼污染屬性字典
        if mc:
            if ff_type_name:
                cur.execute(
                    'INSERT INTO eq_type (group_code, code, name, sort) VALUES (?,?,?,?) '
                    'ON CONFLICT(group_code, code) DO UPDATE SET name=excluded.name',
                    (group_code, type_code, ff_type_name, int(type_code) if type_code.isdigit() else 0))
            if ff_attr_name and attr_code:
                key = (group_code, type_code, attr_code)
                if key in attr_owner and attr_owner[key] != ff_attr_name:
                    msg = (f'{code}：屬性碼 {group_code}{type_code}-{attr_code} 已是'
                           f'「{attr_owner[key]}」，本列卻是「{ff_attr_name}」')
                    warnings.append(msg)
                    fix_reasons.append(f'加工屬性碼撞號（已被「{attr_owner[key]}」使用）')
                else:
                    attr_owner[key] = ff_attr_name
                    cur.execute(
                        'INSERT INTO eq_attr (group_code, type_code, code, name) VALUES (?,?,?,?) '
                        'ON CONFLICT(group_code, type_code, code) DO UPDATE SET name=excluded.name',
                        (group_code, type_code, attr_code, ff_attr_name))

        remark, note = c[COL_REMARK], c[COL_NOTE]
        row_vals = (group_code, type_code, attr_code, seq, c[COL_OLD], c[COL_VENDOR],
                    c[COL_BUY], remark, note, ff_type_name, ff_attr_name, c[COL_LOCATION],
                    guess_status(note, remark),
                    1 if fix_reasons else 0, '；'.join(fix_reasons))

        src = existing.get(code)
        if src == 'manual':
            skipped += 1                                # 系統內新增的設備不讓 Excel 蓋掉
            continue
        if src is None:
            cur.execute(
                'INSERT INTO equipment (code, group_code, type_code, attr_code, seq, old_code, '
                'vendor, buy_date, remark, note, type_name_raw, attr_name_raw, location, status, '
                'needs_fix, fix_reason, source, origin) '
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'excel','excel')", (code,) + row_vals)
            added += 1
        else:
            cur.execute(
                'UPDATE equipment SET group_code=?, type_code=?, attr_code=?, seq=?, old_code=?, '
                'vendor=?, buy_date=?, remark=?, note=?, type_name_raw=?, attr_name_raw=?, '
                'location=?, status=?, needs_fix=?, '
                "fix_reason=?, updated_at=datetime('now','localtime') WHERE code=?",
                row_vals + (code,))
            updated += 1
        if fix_reasons:
            fixcnt += 1
        seen_codes.append(code)

        # 規格與移轉歷程是從 Excel 推導出來的，每次匯入重算（不動人工另外加的）
        cur.execute("DELETE FROM eq_spec WHERE code=?", (code,))
        for name, value, sort in parse_specs(remark):
            cur.execute('INSERT INTO eq_spec (code, spec_name, spec_value, sort) VALUES (?,?,?,?)',
                        (code, name, value, sort))
        cur.execute("DELETE FROM eq_history WHERE code=? AND user='excel'", (code,))
        for date, action, detail in parse_move_history(c[COL_MOVE]):
            cur.execute('INSERT INTO eq_history (code, date, action, detail, user) '
                        "VALUES (?,?,?,?,'excel')", (code, date, action, detail))

    for gcode, gname in group_names.items():
        cur.execute('INSERT INTO eq_group (code, name, sort) VALUES (?,?,?) '
                    'ON CONFLICT(code) DO UPDATE SET name=excluded.name',
                    (gcode, gname, ord(gcode)))
    conn.commit()
    wb.close()

    # Excel 裡不見了的設備不刪除（可能是別人手滑刪的），只提示
    gone = [c for c, s in existing.items() if s == 'excel' and c not in seen_codes]
    if gone:
        warnings.append(f'DB 有但本次 Excel 沒有的設備 {len(gone)} 筆（未刪除）：{", ".join(gone[:10])}')

    # 舊編號重複提示（砂輪機共用 B10 屬正常，車06 之類才需要確認）
    dup = cur.execute(
        'SELECT old_code, COUNT(*) c, GROUP_CONCAT(code) FROM equipment '
        "WHERE old_code<>'' GROUP BY old_code HAVING c>1").fetchall()
    for old, cnt, codes in dup:
        warnings.append(f'舊編號「{old}」被 {cnt} 台共用：{codes}')

    return added, updated, skipped, fixcnt, warnings


def _scan_folder(root, want_ext=None):
    """走訪一層層資料夾，回傳 (relpath, 頂層資料夾名, 檔名, size, mtime)"""
    out = []
    if not os.path.isdir(root):
        print(f'  [WARN] 資料夾不存在或無法存取：{root}')
        return out
    for dirpath, dirnames, filenames in os.walk(root):
        for fn in filenames:
            if fn.lower() in SKIP_FILES or fn.startswith('~$'):
                continue
            ext = os.path.splitext(fn)[1].lower()
            if want_ext and ext not in want_ext:
                continue
            full = os.path.join(dirpath, fn)
            rel = os.path.relpath(full, root).replace(os.sep, '/')
            top = rel.split('/')[0] if '/' in rel else ''
            try:
                st = os.stat(full)
            except OSError:
                continue
            out.append((rel, top, fn, st.st_size,
                        datetime.datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S')))
    return out


def scan_photos(conn):
    """掃描設備照片資料夾，資料夾名開頭是 PDM 編碼者自動歸位，其餘留 code=NULL 待認領"""
    root = config.EQUIPMENT_PHOTO_ROOT
    print(f'掃描設備照片：{root}')
    files = _scan_folder(root, PHOTO_EXT)
    cur = conn.cursor()
    # 人工認領過的對應關係要留著，不能被自動掃描洗掉
    claimed = {r[0]: (r[1], r[2]) for r in cur.execute(
        'SELECT relpath, code, is_cover FROM eq_photo WHERE claimed=1')}
    valid = {r[0] for r in cur.execute('SELECT code FROM equipment')}

    cur.execute('DELETE FROM eq_photo')
    matched = orphan = 0
    for rel, top, fn, size, mtime in files:
        m = FOLDER_CODE_RE.match(top)
        code = m.group(1) if m and m.group(1) in valid else None
        is_cover, is_claimed = 0, 0
        if rel in claimed:
            code, is_cover = claimed[rel]
            is_claimed = 1
        cur.execute('INSERT OR REPLACE INTO eq_photo '
                    '(relpath, code, folder, filename, size, mtime, is_cover, claimed) '
                    'VALUES (?,?,?,?,?,?,?,?)', (rel, code, top, fn, size, mtime, is_cover, is_claimed))
        if code:
            matched += 1
        else:
            orphan += 1

    # 每台設備挑一張當封面（同資料夾內檔名排序第一張）
    cur.execute('UPDATE eq_photo SET is_cover=0 WHERE claimed=0')
    cur.execute("""UPDATE eq_photo SET is_cover=1 WHERE relpath IN (
                     SELECT MIN(relpath) FROM eq_photo WHERE code IS NOT NULL GROUP BY code)""")
    conn.commit()
    return matched, orphan


def scan_tech_files(conn):
    """掃描設備技術資料資料夾，同樣以編碼前綴歸位"""
    root = config.EQUIPMENT_TECH_ROOT
    print(f'掃描技術資料：{root}')
    files = _scan_folder(root)
    cur = conn.cursor()
    valid = {r[0] for r in cur.execute('SELECT code FROM equipment')}
    cur.execute('DELETE FROM eq_tech_file')
    matched = 0
    for rel, top, fn, size, mtime in files:
        m = FOLDER_CODE_RE.match(top)
        code = m.group(1) if m and m.group(1) in valid else None
        cur.execute('INSERT OR REPLACE INTO eq_tech_file '
                    '(relpath, code, folder, filename, ext, size, mtime) VALUES (?,?,?,?,?,?,?)',
                    (rel, code, top, fn, os.path.splitext(fn)[1].lower(), size, mtime))
        if code:
            matched += 1
    conn.commit()
    return matched, len(files) - matched


def _deploy(src):
    dist_app = os.path.join(_APP_DIR, 'dist_embed', 'PDS系統', '_app')
    if not os.path.isdir(dist_app):
        print(f'  [WARN] dist_embed 目錄不存在，略過部署：{dist_app}')
        return
    shutil.copy2(src, os.path.join(dist_app, os.path.basename(src)))
    print(f'  部署 {os.path.basename(src)} -> {dist_app}')


def _migrate(conn):
    """相容遷移：舊版 equipment.db 缺新欄位時補上"""
    existing = {r[1] for r in conn.execute('PRAGMA table_info(equipment)')}
    for col in ('type_name_raw', 'attr_name_raw'):
        if col not in existing:
            conn.execute(f'ALTER TABLE equipment ADD COLUMN {col} TEXT')
    if 'origin' not in existing:
        conn.execute("ALTER TABLE equipment ADD COLUMN origin TEXT DEFAULT 'excel'")
        # 舊資料沒有 origin：source=manual 的當初就是系統內新增的，其餘來自 Excel
        conn.execute("UPDATE equipment SET origin = source WHERE origin IS NULL OR origin=''")
    conn.commit()


def rebuild(deploy=False, scan=True):
    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)
    _migrate(conn)

    result = import_excel(conn)
    if result is None:
        conn.close()
        return -1
    added, updated, skipped, fixcnt, warnings = result
    print(f'  設備：新增 {added}、更新 {updated}、略過(系統內新增) {skipped}、待修正 {fixcnt}')

    if scan:
        pm, po = scan_photos(conn)
        print(f'  照片：已歸位 {pm} 張、待認領 {po} 張')
        tm, to = scan_tech_files(conn)
        print(f'  技術資料：已歸位 {tm} 個、未對應 {to} 個')

    total = conn.execute('SELECT COUNT(*) FROM equipment').fetchone()[0]
    groups = conn.execute('SELECT COUNT(*) FROM eq_group').fetchone()[0]
    types = conn.execute('SELECT COUNT(*) FROM eq_type').fetchone()[0]
    conn.close()

    if warnings:
        print(f'\n提醒（{len(warnings)} 則，不影響匯入）：')
        for w in warnings:
            print(f'  - {w}')

    print(f'\n完成！設備 {total} 台、群組 {groups} 個、機台類型 {types} 種')
    print(f'DB -> {DB_PATH}')
    if deploy:
        _deploy(DB_PATH)
    return total


if __name__ == '__main__':
    args = sys.argv[1:]
    start = datetime.datetime.now()
    print('=' * 55)
    print('  設備主檔匯入工具')
    print(f'  時間: {start:%Y-%m-%d %H:%M:%S}')
    print('=' * 55)

    result = rebuild(deploy='--deploy' in args, scan='--no-scan' not in args)

    print(f'\n總耗時: {(datetime.datetime.now() - start).total_seconds():.1f} 秒')
    if result < 0:
        sys.exit(1)
