#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
油品主檔匯入／索引重建工具

掃描 config.OIL_MSDS_ROOT 的 MSDS PDF 檔名，解析出「供應商／油品代號／品名」建立油品主檔，
並掃描 config.OIL_DOC_ROOT 的簡介文件、匯入舊的「油品更換記錄表.xlsx」。

設計重點（詳見 docs/oil-management.md）：
  * oil.db 是**正本**，MSDS 檔名只是第一次的匯入來源。匯入採合併（merge）：
    - source='msds' 的油品才會被檔名覆蓋
    - source='manual'（系統內新增或改過的）永遠不會被覆蓋或刪除
    - 檔案消失的油品不會被刪掉，只在報表提示
  * MSDS\報廢\ 底下的檔案 → 該油品狀態預設「停用」（但只在第一次建立時套用，
    之後使用者在系統內改的狀態最大，不會被重新掃描蓋回去）
  * 檔案索引（oil_file）每次掃描重建，人工認領的對應（claimed=1）保留

用法：
  python build_oil_index.py                # 掃 MSDS/簡介 + 匯入更換記錄表
  python build_oil_index.py --no-change    # 不匯入更換記錄表（快）
  python build_oil_index.py --deploy       # 相容用旗標，oil.db 在網芳，不需要複製
"""

import os
import re
import sys
import sqlite3
import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

_APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = getattr(config, 'OIL_DB_PATH', None) or os.path.join(_APP_DIR, 'oil.db')

MSDS_ROOT = getattr(config, 'OIL_MSDS_ROOT', '')
DOC_ROOT = getattr(config, 'OIL_DOC_ROOT', '')
CHANGE_XLSX = getattr(config, 'OIL_CHANGE_LEGACY_XLSX', '')
EQ_DB_PATH = getattr(config, 'EQUIPMENT_DB_PATH', None) or os.path.join(_APP_DIR, 'equipment.db')

# 檔名裡的中括號欄位：[唯勝][AW-30]STORK金屬加工液 SW5031-030_2025.1.9.pdf
BRACKET_RE = re.compile(r'\[([^\]]+)\]')
# 檔名尾端的版本日期：_2025.1.9 / _2025.1.20
FILE_DATE_RE = re.compile(r'_(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})\s*$')
# 「被DS32取代」「升級成9520」「S7取代」這類接手油品的線索
REPLACED_RE = re.compile(r'[被,，]?\s*([A-Za-z0-9\-]{2,10})\s*(?:取代|接手)')
UPGRADE_RE = re.compile(r'(?:升級成|改用)\s*([A-Za-z0-9\-]{2,10})')

# 中括號第一欄是供應商還是油品代號，靠這份名單分辨（見 docs/oil-management.md）
KNOWN_SUPPLIERS = {'唯勝', '美科', '昱礽', '久昌', '快密刀'}

# 檔名裡沒有情報量的字（SDS/MSDS/GHS/安全標示…），清掉後剩不到東西就用代號當品名
NAME_NOISE_RE = re.compile(r'^(?:SDS|MSDS|GHS|安全標示|安全資料表)[\s_\-()（）]*', re.I)

# 停用（報廢）油品放在 MSDS\報廢\；2025.3.5暫存 是同一批檔案的副本，整個略過
OBSOLETE_DIRS = {'報廢'}
SKIP_DIRS = {'2025.3.5暫存'}

SKIP_FILES = {'thumbs.db', 'desktop.ini'}
DOC_EXT = {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.xlsm', '.ppt', '.pptx', '.txt',
           '.jpg', '.jpeg', '.png'}

# 品名關鍵字 → 分類（由上往下比對，先中先算，所以「金屬加工液」要排在「加工油」前面）
CATEGORY_RULES = [
    ('切削液', '水性切削液'), ('金屬加工液', '水性切削液'), ('加工液', '水性切削液'),
    ('油膏', '切削油膏'),
    ('金屬加工油', '切削油'), ('切削油', '切削油'),
    ('液壓油', '液壓油'),
    ('拖板油', '滑道油'), ('滑道油', '滑道油'), ('滑軌油', '滑道油'),
    ('主軸油', '主軸油'),
    ('齒輪油', '齒輪油'),
    ('防銹油', '防銹油'), ('防鏽油', '防銹油'),
    ('清潔劑', '清潔劑'), ('清洗劑', '清潔劑'),
    ('煤油', '燃料'), ('柴油', '燃料'), ('汽油', '燃料'),
    ('機油', '機油'), ('潤滑脂', '潤滑脂'), ('黃油', '潤滑脂'),
]

SCHEMA = """
CREATE TABLE IF NOT EXISTS oil (
    code        TEXT PRIMARY KEY,      -- 油品代號：AW-30 / DS-32 / 9520 / 煤油
    name        TEXT,                  -- 品名
    brand       TEXT,                  -- 品牌（STORK / SUPERSYN…）
    supplier    TEXT,                  -- 供應商
    category    TEXT,                  -- 分類（水性切削液/液壓油/滑道油…）
    spec        TEXT,                  -- 規格（黏度、稀釋濃度…）
    pack        TEXT,                  -- 包裝（18L桶 / 200L鐵桶）
    status      TEXT DEFAULT '使用中',  -- 使用中 / 停用
    replaced_by TEXT,                  -- 停用後由哪支油品接手
    usage_note  TEXT,                  -- 使用說明／注意事項
    remark      TEXT,
    -- source / origin 的分工同設備主檔（見 docs/equipment-master.md）：
    --   source = 這筆還能不能被重新掃描覆蓋（系統內改過就變 manual）
    --   origin = 這筆當初打哪來的，寫入後永不變更，用來擋「掃描來源的油品被硬刪」
    source      TEXT DEFAULT 'msds',
    origin      TEXT DEFAULT 'msds',
    created_at  TEXT DEFAULT (datetime('now','localtime')),
    updated_at  TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS oil_file (
    relpath  TEXT PRIMARY KEY,         -- 相對 root 的路徑
    root     TEXT,                     -- 'msds' | 'doc'（決定用哪個根目錄還原絕對路徑）
    code     TEXT,                     -- NULL = 尚未歸位（檔名看不出代號）
    folder   TEXT, filename TEXT, ext TEXT,
    size     INTEGER, mtime TEXT,
    file_date TEXT,                    -- 檔名解析出的版本日期
    obsolete INTEGER DEFAULT 0,        -- 1 = 放在「報廢」資料夾
    claimed  INTEGER DEFAULT 0         -- 1 = 人工認領/上傳的對應，重新掃描時保留
);
CREATE TABLE IF NOT EXISTS oil_unit (
    code TEXT, unit TEXT, equip_code TEXT, note TEXT, sort INTEGER
);
CREATE TABLE IF NOT EXISTS oil_change (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT, equip_code TEXT, equip_name TEXT,
    date TEXT, qty TEXT, operator TEXT, note TEXT,
    user TEXT DEFAULT 'user',          -- excel = 舊表匯入、user = 系統內登錄
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS oil_history (
    code TEXT, date TEXT, action TEXT, detail TEXT, user TEXT
);
CREATE INDEX IF NOT EXISTS idx_oil_file_code ON oil_file(code);
CREATE INDEX IF NOT EXISTS idx_oil_unit_code ON oil_unit(code);
CREATE INDEX IF NOT EXISTS idx_oil_chg_code  ON oil_change(code);
CREATE INDEX IF NOT EXISTS idx_oil_chg_eq    ON oil_change(equip_code);
CREATE INDEX IF NOT EXISTS idx_oil_hist_code ON oil_history(code);
"""


def _now():
    return datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def norm_code(s):
    """代號比對用的正規化：去掉連字號/空白、英文轉大寫（AW-30 == AW30 == aw 30）"""
    return re.sub(r'[\s\-_.]', '', (s or '')).upper()


def guess_category(text):
    for kw, cat in CATEGORY_RULES:
        if kw in text:
            return cat
    return '其他'


def guess_brand(text):
    for b in ('STORK', 'SUPERSYN', 'HYSLIP', 'MCM'):
        if b.lower() in text.lower():
            return b
    return ''


def _clean_name(text):
    """清掉檔名裡的文件類型字樣與純數字流水號，剩下的才是真正的品名。

    `[煤油] GHS_2025.1.20.pdf` 解析出來的「GHS」不是品名，清掉後回傳空字串，
    呼叫端就會退而用代號當品名（顯示「煤油」而不是「GHS」）。"""
    s = (text or '').strip(' -_,')
    while True:
        t = NAME_NOISE_RE.sub('', s).strip(' -_,')
        if t == s:
            break
        s = t
    # 清完只剩數字/日期/符號（SDS1080225 → 1080225）就當作沒有品名
    if not re.search(r'[A-Za-z一-鿿]', s):
        return ''
    return s


def parse_msds_name(filename):
    """把 MSDS 檔名拆成 {code, supplier, name, brand, category, file_date, replaced_by}

    檔名格式（實際觀察到的四種）：
      [唯勝][AW-30]STORK金屬加工液 SW5031-030_2025.1.9.pdf   → 供應商+代號
      [美科][9520]SDS SUPERSYN 9520詠基1150423-3.pdf         → 供應商+代號
      [油污清潔劑]MSDS.pdf                                    → 只有代號
      [唯勝]SE0551-2 STORK抗磨損液壓油 AW 32,被DS32取代.pdf    → 只有供應商（報廢區常見）

    第一個中括號是供應商還是代號，用 KNOWN_SUPPLIERS 名單判斷；兩者都判不出來時
    退而用括號外的文字開頭當代號，讓它至少能進系統被看到（使用者再改）。
    """
    base = os.path.splitext(filename)[0]
    tags = BRACKET_RE.findall(base)
    rest = BRACKET_RE.sub('', base).strip(' -_')

    supplier, code = '', ''
    if tags:
        if tags[0] in KNOWN_SUPPLIERS:
            supplier = tags[0]
            if len(tags) > 1:
                code = tags[1]
        else:
            code = tags[0]
            if len(tags) > 1 and not supplier:
                supplier = tags[1] if tags[1] in KNOWN_SUPPLIERS else supplier

    file_date = ''
    m = FILE_DATE_RE.search(rest)
    if m:
        file_date = '%04d-%02d-%02d' % (int(m.group(1)), int(m.group(2)), int(m.group(3)))
        rest = rest[:m.start()].strip()

    if not code:
        # 只有供應商的舊檔（報廢區）：拿括號外的第一段當代號，例如 SE0551-2、CS-1010
        code = (re.split(r'[\s_]+', rest) or [''])[0].strip(' -_,')

    name = _clean_name(rest) or code
    replaced_by = ''
    for rx in (REPLACED_RE, UPGRADE_RE):
        m = rx.search(base)
        if m:
            replaced_by = m.group(1).strip()
            break

    return {
        'code': code.strip(),
        'supplier': supplier,
        'name': name,
        'brand': guess_brand(base),
        'category': guess_category(base),
        'file_date': file_date,
        'replaced_by': replaced_by,
    }


def walk_files(root, skip_dirs):
    """列出資料夾底下所有檔案，回傳 (相對路徑, 絕對路徑, 第一層子資料夾名稱)"""
    out = []
    if not root or not os.path.isdir(root):
        return out
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in skip_dirs]
        for fn in filenames:
            if fn.lower() in SKIP_FILES:
                continue
            full = os.path.join(dirpath, fn)
            rel = os.path.relpath(full, root)
            parts = rel.split(os.sep)
            folder = parts[0] if len(parts) > 1 else ''
            out.append((rel, full, folder))
    return out


def scan_msds(conn):
    """掃 MSDS 資料夾：建立/更新油品主檔 + 重建 MSDS 檔案索引"""
    files = walk_files(MSDS_ROOT, SKIP_DIRS)
    if not files and not os.path.isdir(MSDS_ROOT):
        print(f'  [WARN] MSDS 資料夾不存在：{MSDS_ROOT}')
        return 0, 0

    cur = conn.cursor()
    existing = {r['code']: dict(r) for r in cur.execute('SELECT * FROM oil')}
    claimed = {r['relpath']: r['code'] for r in
               cur.execute("SELECT relpath, code FROM oil_file WHERE claimed=1")}
    cur.execute("DELETE FROM oil_file WHERE root='msds' AND claimed=0")

    new_oil = upd_oil = 0
    for rel, full, folder in files:
        ext = os.path.splitext(rel)[1].lower()
        if ext not in DOC_EXT:
            continue
        info = parse_msds_name(os.path.basename(rel))
        obsolete = 1 if folder in OBSOLETE_DIRS else 0
        code = claimed.get(rel) or info['code']
        if not code:
            code = None

        st = os.stat(full)
        cur.execute("""INSERT OR REPLACE INTO oil_file
            (relpath, root, code, folder, filename, ext, size, mtime, file_date, obsolete, claimed)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (rel, 'msds', code, folder, os.path.basename(rel), ext, st.st_size,
             datetime.datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
             info['file_date'], obsolete, 1 if rel in claimed else 0))

        if not code:
            continue
        row = existing.get(code)
        if row is None:
            # 報廢資料夾裡的油品第一次建立時就標「停用」，之後使用者改了不會被蓋回去
            cur.execute("""INSERT INTO oil
                (code, name, brand, supplier, category, status, replaced_by,
                 source, origin, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,'msds','msds',?,?)""",
                (code, info['name'], info['brand'], info['supplier'], info['category'],
                 '停用' if obsolete else '使用中', info['replaced_by'], _now(), _now()))
            existing[code] = {'code': code, 'source': 'msds'}
            new_oil += 1
        elif row.get('source') != 'manual' and not obsolete:
            # 掃描來源的資料才更新；已經在系統內編輯過的（manual）完全不動。
            # 「報廢」資料夾裡的舊版檔案不參與更新（它的檔名常帶「,被DS32取代」這種
            # 註記，拿去覆蓋現行品名只會把好資料弄髒），只有 replaced_by 值得留下——
            # 那是在建立記錄時就寫進去的。
            # 停用中的油品在 MSDS 根目錄又出現一份新檔案時，也不自動改回使用中——
            # 「這支油還在不在用」是人的決策，不該由檔案位置反推
            cur.execute("""UPDATE oil SET name=?, brand=?, supplier=?, category=?,
                                          replaced_by=COALESCE(NULLIF(?,''), replaced_by),
                                          updated_at=?
                            WHERE code=?""",
                        (info['name'] or row.get('name'), info['brand'] or row.get('brand'),
                         info['supplier'] or row.get('supplier'),
                         info['category'] or row.get('category'),
                         info['replaced_by'], _now(), code))
            upd_oil += 1
    conn.commit()
    return new_oil, upd_oil


def scan_docs(conn):
    """掃 油品簡介 資料夾：純檔案索引，靠檔名/資料夾名裡的代號歸位（找不到就留空）"""
    files = walk_files(DOC_ROOT, SKIP_DIRS)
    cur = conn.cursor()
    codes = [r['code'] for r in cur.execute('SELECT code FROM oil')]
    ncodes = sorted(((norm_code(c), c) for c in codes), key=lambda x: -len(x[0]))
    claimed = {r['relpath']: r['code'] for r in
               cur.execute("SELECT relpath, code FROM oil_file WHERE claimed=1")}
    cur.execute("DELETE FROM oil_file WHERE root='doc' AND claimed=0")

    n = 0
    for rel, full, folder in files:
        ext = os.path.splitext(rel)[1].lower()
        if ext not in DOC_EXT:
            continue
        code = claimed.get(rel)
        if not code:
            hay = norm_code(rel)
            # 長的代號先比，避免「S9」誤命中「9520」這種包含關係
            for nc, orig in ncodes:
                if len(nc) >= 3 and nc in hay:
                    code = orig
                    break
        st = os.stat(full)
        cur.execute("""INSERT OR REPLACE INTO oil_file
            (relpath, root, code, folder, filename, ext, size, mtime, file_date, obsolete, claimed)
            VALUES (?,?,?,?,?,?,?,?,'',0,?)""",
            (rel, 'doc', code, folder, os.path.basename(rel), ext, st.st_size,
             datetime.datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
             1 if rel in claimed else 0))
        n += 1
    conn.commit()
    return n


def _equipment_old_map():
    """舊設備編號（綜銑01…）→ PDM 編碼。設備主檔讀不到時回傳空 dict，不影響匯入。"""
    out = {}
    try:
        if not os.path.exists(EQ_DB_PATH):
            return out
        eq = sqlite3.connect(EQ_DB_PATH, timeout=15)
        for code, old in eq.execute("SELECT code, IFNULL(old_code,'') FROM equipment"):
            for tok in re.split(r'[,、/;\s]+', old):
                tok = tok.strip()
                if tok and tok not in out:
                    out[tok] = code
        eq.close()
    except Exception as exc:
        print(f'  [WARN] 讀取設備主檔失敗（更換記錄不對應設備）：{exc}')
    return out


def import_change_xlsx(conn):
    """匯入舊的「油品更換記錄表.xlsx」

    表格長相：A1 標題（【加工課】油品記錄表(水性) AW30）→ 決定這張表是哪支油品；
    第 3 列表頭（設備 / 1 / 2 / 3…）；第 4 列起每列一台設備，A=設備舊編號、
    B=機台型號、C 欄以後是第 1、2、3… 次的更換日期。

    重複匯入時先清掉 user='excel' 的舊資料再寫入，系統內登錄的（user='user'）不動。
    """
    if not CHANGE_XLSX or not os.path.exists(CHANGE_XLSX):
        print(f'  [SKIP] 更換記錄表不存在：{CHANGE_XLSX}')
        return 0
    try:
        import openpyxl
    except ImportError:
        print('  [SKIP] 未安裝 openpyxl，略過更換記錄匯入')
        return 0

    wb = openpyxl.load_workbook(CHANGE_XLSX, data_only=True)
    cur = conn.cursor()
    oil_codes = {norm_code(r['code']): r['code'] for r in cur.execute('SELECT code FROM oil')}
    old_map = _equipment_old_map()
    cur.execute("DELETE FROM oil_change WHERE user='excel'")

    total = 0
    for ws in wb.worksheets:
        title = ' '.join(str(c.value) for c in ws[1] if c.value) if ws.max_row >= 1 else ''
        # 標題末尾的代號（AW30）對回主檔的 AW-30
        code = ''
        for tok in re.findall(r'[A-Za-z0-9\-]{2,12}', title):
            if norm_code(tok) in oil_codes:
                code = oil_codes[norm_code(tok)]
                break
        if not code:
            print(f'  [WARN] 分頁「{ws.title}」標題認不出油品代號，略過：{title[:40]}')
            continue

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue
            equip_name = str(row[0]).strip()
            model = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            equip_code = old_map.get(equip_name, '')
            for v in row[2:]:
                if v is None or str(v).strip() == '':
                    continue
                if isinstance(v, datetime.datetime):
                    d = v.strftime('%Y-%m-%d')
                elif isinstance(v, datetime.date):
                    d = v.strftime('%Y-%m-%d')
                else:
                    d = str(v).strip()
                cur.execute("""INSERT INTO oil_change
                    (code, equip_code, equip_name, date, qty, operator, note, user)
                    VALUES (?,?,?,?, '', '', ?, 'excel')""",
                    (code, equip_code, equip_name, d, ('機台型號 ' + model) if model else ''))
                total += 1
    conn.commit()
    return total


def main():
    args = set(sys.argv[1:])
    print('=' * 60)
    print('油品主檔匯入')
    print(f'  資料庫：{DB_PATH}')
    print(f'  MSDS  ：{MSDS_ROOT}')
    print('=' * 60)

    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.row_factory = sqlite3.Row
    conn.executescript(SCHEMA)

    new_oil, upd_oil = scan_msds(conn)
    print(f'  [OK] MSDS 掃描完成：新增 {new_oil} 支油品、更新 {upd_oil} 支')
    ndoc = scan_docs(conn)
    print(f'  [OK] 簡介文件索引 {ndoc} 個檔案')

    if '--no-change' not in args:
        nchg = import_change_xlsx(conn)
        print(f'  [OK] 更換記錄匯入 {nchg} 筆')

    total = conn.execute('SELECT COUNT(*) FROM oil').fetchone()[0]
    use = conn.execute("SELECT COUNT(*) FROM oil WHERE status='使用中'").fetchone()[0]
    nomsds = conn.execute(
        "SELECT COUNT(*) FROM oil o WHERE NOT EXISTS "
        "(SELECT 1 FROM oil_file f WHERE f.code=o.code AND f.root='msds')").fetchone()[0]
    orphan = conn.execute("SELECT COUNT(*) FROM oil_file WHERE code IS NULL").fetchone()[0]
    conn.close()

    print('-' * 60)
    print(f'  油品共 {total} 支（使用中 {use}）；缺 MSDS {nomsds} 支；未歸位檔案 {orphan} 個')
    if '--deploy' in args:
        print('  [INFO] oil.db 放在網芳共用，--deploy 不需要做任何複製')
    print('完成。')


if __name__ == '__main__':
    main()
